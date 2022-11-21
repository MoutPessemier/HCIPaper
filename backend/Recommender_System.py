
import pandas as pd
from openpyxl import load_workbook


class Databank:

    def __init__(self):
        self.__recommender_data = pd.DataFrame(columns=['dog_id', 'name', 'sex', 'size', 'age', 'dogfriendly', 'petfriendly','childfriendly','gardenreq','hug','training','BCD','Description','image','score'])  # dataframe maken; handig voor doorzoeken

    def create_dog(self):

        df = pd.read_csv("DogsData" + '.csv', nrows=10000)  # data inlezen
        for row, dog in df.iterrows():  # .itterrows() geeft een rowindex en data weer in de rij
            self.append_dog_data(row, dog['Naam'], dog['Geslacht'], dog['Groot'], dog['Leeftijd'], #We populeren het systeem met de honden
                                   dog['Andere_hond'],
                                   dog['Andere_dieren'], dog['Kinderen'], dog['Tuin'], dog['Knuffelbeer'],
                                   dog['Training_sport'], dog['BCDp'], dog['Beschrijving'], dog['Image'])

    def append_dog_data(self, id, name, sex, size, age, dogfriendly, petfriendly, childfriendly, gardenreq, hug, training, BCD, Description, image, score= 0):


        new_row = pd.Series({'dog_id': id, 'name': name, 'size': size, 'sex': sex,
             'age': age, 'dogfriendly': dogfriendly, 'petfriendly': petfriendly, 'childfriendly': childfriendly,'gardenreq': gardenreq,
             'hug': hug,'training': training,'BCD': BCD,'score': score,'description': Description, 'image': image})
        self.__recommender_data = pd.concat([self.__recommender_data, new_row.to_frame().T], ignore_index=True) #aangezien een pd.series 1 kolom geeft met de vragen als rij moeten we transponern met .T

    def add_points(self, id, score):
        for row, dog in self.__recommender_data.iterrows():  # .itterrows() geeft een rowindex en data weer in de rij
            if dog['dog_id'] == id:
                self.__recommender_data.loc[row, 'score']+= score

    def make_recommendation(self, inputdata):
        for vraagnum, data in inputdata.iterrows():


            input = data['antwoorden']
            gewicht = data['gewichten']

            for row, dog in self.give_data().iterrows():  # we gaan door elke hond in onze database

                # VRAAG 1: Welk geslacht prefereert u?  (eigenschap: geslacht → reu, teef)
                if vraagnum == 0: #Merk op dat de indices begint bij 0 en NIET 1
                    if dog['sex'] == input:  # V1: We willen absoluut de juiste geslacht honden, daarom 2 als basis
                        self.add_points(dog['dog_id'], 3 + ((gewicht - 50) / 50))
                        if dog['name'] == "r":
                            print('Een punt aan', dog['name'], 'voor vraag', vraagnum+1)

                # VRAAG 2: Welk grootte prefereert u?  (eigenschap: geslacht → reu, teef)
                if vraagnum == 1:
                    if dog['size'] == input:  # V2: We willen ook zeker de juiste grootte, maar minder belangrijk dan geslacht
                        self.add_points(dog['dog_id'], 1 + ((gewicht - 50) / 50))
                        if dog['name'] == "r":
                            print('Een punt aan', dog['name'], 'voor vraag', vraagnum+1)

                # VRAAG 3: Leeftijd?  (eigenschap: geslacht → reu, teef)
                if vraagnum == 2:
                    if dog['age'] == input:  # V2: We willen ook zeker de juiste grootte, maar minder belangrijk dan geslacht
                        self.add_points(dog['dog_id'], 1 + ((gewicht - 50) / 50))
                        if dog['name'] == "r":
                            print('Een punt aan', dog['name'], 'voor vraag', vraagnum+1)

                #VRAAG 4: Heeft u reeds (een) andere hond(en)? (eigenschap: ja, reu, teef, nee)
                #Note to self: te harde codering!
                if vraagnum == 3:
                    if dog['dogfriendly'] == input and input == 'Ja, een reu':
                        self.add_points(dog['dog_id'], 1 + ((gewicht - 50) / 50))
                        if dog['name'] == "r":
                            print('Een punt aan', dog['name'], 'voor vraag', vraagnum+1)
                    if dog['dogfriendly'] == input and input == 'Ja, een teef':
                        self.add_points(dog['dog_id'], 1 + ((gewicht - 50) / 50))
                        if dog['name'] == "r":
                            print('Een punt aan', dog['name'], 'voor vraag', vraagnum+1)
                    if dog['dogfriendly'] == input and input == 'Ja, meerdere honden':
                        self.add_points(dog['dog_id'], 1 + ((gewicht - 50) / 50))
                        if dog['name'] == "r":
                            print('Een punt aan', dog['name'], 'voor vraag', vraagnum+1)
                    if dog['dogfriendly'] == input and input == 'Nee':
                        self.add_points(dog['dog_id'], 1 + ((gewicht - 50) / 50))
                        if dog['name'] == "r":
                            print('Een punt aan', dog['name'], 'voor vraag', vraagnum+1)

                # VRAAG 5 want deze vragen hebben dezelfde puntenlogica en 1/1 matching met de dataset
                if vraagnum == 4:
                    if dog['petfriendly'] == input:  # V3: Gewicht beslissen mensen zelf
                        self.add_points(dog['dog_id'], 1 + ((gewicht - 50) / 50))
                        if dog['name'] == "r":
                            print('Een punt aan', dog['name'], 'voor vraag', vraagnum+1)

#-------- Vanaf hier speciale logica vragen:
                # Note to self: te harde codering!
                # VRAAG 6: Wat voor gezinssituatie heeft u? (eigenschap: ja, grote, nee)
                if vraagnum == 5:
                    if dog['childfriendly'] == input and input == "Enkel kinderen boven 12 jaar":
                        self.add_points(dog['dog_id'], 1 + ((gewicht - 50) / 50))
                        if dog['name'] == "r":
                            print('Een punt aan', dog['name'], 'voor vraag', vraagnum+1)
                    else:  # V6 als er dus enkel volwassene zijn krijgt iedereen een punt
                        #self.give_all()
                        continue


                # --------
                # VRAAG 7: Heeft u een omheinde tuin? (eigenschap: default, moet tuin: ja)
                if vraagnum == 6:
                    if dog['gardenreq'] == input and input == "Nee":  # V7, indien input "nee" dan enkel honden met
                        self.add_points(dog['dog_id'], 1 + ((gewicht - 50) / 50))
                        if dog['name'] == "r":
                            print('Een punt aan', dog['name'], 'voor vraag', vraagnum+1)
                    if dog['gardenreq'] == input and input == "Ja":  # V7, indien input "ja" dan alle honden een punt, equivalent als zeggen dat niemand iets krijgt.
                        #self.give_all()
                        continue


                # --------
                # VRAAG 8: Zoekt u een echte knuffelbeer? (eigenschap: default, ja)
                if vraagnum == 7:
                    if dog['hug'] == input and input == "Ja, knuffelkontjes!":  # V8
                        self.add_points(dog['dog_id'], 1 + ((gewicht - 50) / 50))
                        if dog['name'] == "r":
                            print('Een punt aan', dog['name'], 'voor vraag', vraagnum+1)
                    # else:
                    #     self.give_all()

                # --------
                # VRAAG 9: Zou u graag training en/of hondensport doen met de hond? (eigenschap: default, training/sport)
                if vraagnum == 8:
                    if dog['training'] == input and input == "Ja, graag!":  # V9
                        self.add_points(dog['dog_id'], 1 + ((gewicht - 50) / 50))
                        if dog['name'] == "r":
                            print('Een punt aan', dog['name'], 'voor vraag', vraagnum+1)
                    # else:
                    #     self.give_all()

                # --------
                #VRAAG 10: Prefereert u een hond die getraind wordt via het Belgian Cell Dogs project?* (eigenschap: default, BCDp)
                if vraagnum == 9:
                    if dog['BCD'] == input and input == "Ja":  # V10
                        self.add_points(dog['dog_id'], 1 + ((gewicht - 50) / 50))
                        if dog['name'] == "r":
                            print('Een punt aan', dog['name'], 'voor vraag', vraagnum+1)
                    # else:
                    #     self.give_all()
        return self.give_top4()
    def give_all(self):
        for row, dog in self.__recommender_data.iterrows():  # .itterrows() geeft een rowindex en data weer in de row
            self.__recommender_data.loc[row, 'score']+= 1/len(self.__recommender_data) #we normalizeren tegenover het aantal honden in de lijst

    def give_data(self):
        return self.__recommender_data

    def search_dupl(self):
        for row1, dog1 in self.__recommender_data.iterrows():  # .itterrows() geeft een rowindex en data weer in de rij
            count =1
            for row2, dog2 in self.__recommender_data.iterrows():
                if row1==row2:
                    break
                else:
                    if dog1['name']== dog2['name']:
                        self.__recommender_data.loc[row1, 'score'] = self.__recommender_data.loc[row2, 'score']

    def give_top4(self):
        max_value = 0
        self.search_dupl() #om honden die samengezet moeten worden dezelfde punten gaan krijgen, assumptie: enkel duplicates in data indien ze samen horen
        for row, data in self.__recommender_data.iterrows():
            if max_value < self.__recommender_data.loc[row, 'score']:
                max_value = self.__recommender_data.loc[row, 'score']
            if max_value == 0:
                max_value+=1

        for row, data in self.__recommender_data.iterrows():
            #print(max_value)
            x = 10 * (self.__recommender_data.loc[row, 'score'] / max_value) #als we een schaal op 10 willen is deze normalizatie noodzakelijk omdat de gewichten zorgen voor onbalans
            #x = self.__recommender_data.loc[row, 'score']
            self.__recommender_data.loc[row, 'score'] = round(x, 3) #want 10 is de schaal waarmee we werken

        lijst = []
        for row, data in self.__recommender_data.sort_values(by=['score', 'name'], ascending=False).head(4).iterrows():
            dicti = {}
            dicti["name"] = data['name']
            dicti["imgURL"] = data['image']
            dicti["description"] = data['description']
            lijst.append(dicti)

        return lijst
        #return self.__recommender_data.sort_values(by=['score', 'name'], ascending=False).head(4)

    def export_excel(self, lijst):
        workbook = load_workbook('/Users/nam/PycharmProjects/HCIPaper/backend/records.xlsx')  # je laadt een workbook
        ws = workbook.active  # je activeert e workbook
        identificatie= id(lijst) #elke run zal er een andere id gegeneert worden obv van dit getal
        inzet = [identificatie] #meer leren of IDENT ID
        for dicts in lijst:
            for key, value in dicts.items():
                inzet.append(value)

        ws.append(inzet)
        print('appended succesfully')
        workbook.template = False
        workbook.save('/Users/nam/PycharmProjects/HCIPaper/backend/records.xlsx')
        print('saved')
        return identificatie

    def read_excel(self, id):
        workbook = load_workbook('records.xlsx')  # je laadt een workbook
        ws = workbook.active  # je activeert e workbook
        lijst=[]
        for row in ws.iter_rows(min_row=2, max_col= 13):
            if row[0].value == id: #Dus de rij met het specifieke ID maar de id moet niet terug meegegeven worden
                for x in range(0,10):
                    if x==0 or x==3 or x==6 or x==9:
                        dicti={}
                        dicti["name"] = row[1+x].value
                        #print(dicti["name"])
                        dicti["imgURL"] = row[2+x].value
                        #print(dicti["imgURL"])
                        dicti["description"] = row[3+x].value
                        #print(dicti["description"])
                        #print('-------')
                        lijst.append(dicti)
                        lijst
        return lijst


