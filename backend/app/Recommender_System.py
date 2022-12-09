
import pandas as pd
from openpyxl import load_workbook

class Databank:

    def __init__(self):
        self.__recommender_data = pd.DataFrame(columns=['dog_id', 'name', 'sex', 'size', 'age', 'dogfriendly', 'petfriendly','childfriendly','gardenreq','hug','training','BCD','Description','image','summary','link','score'])

    def create_dog(self):

        df = pd.read_excel("Dogs" + '.xlsx')
        for row, dog in df.iterrows():
            self.append_dog_data(row, dog['Naam'], dog['Geslacht'], dog['Groot'], dog['Leeftijd'], dog['Andere_hond'], dog['Andere_dieren'], dog['Kinderen'], dog['Tuin'], dog['Knuffelbeer'], dog['Training_sport'], dog['BCDp'], dog['Beschrijving'], dog['Image'], dog['Summary'], dog['Link'])

    def append_dog_data(self, id, name, sex, size, age, dogfriendly, petfriendly, childfriendly, gardenreq, hug, training, BCD, Description, image, summary, links, score= 0):

        new_row = pd.Series({'dog_id': id, 'name': name, 'size': size, 'sex': sex,
             'age': age, 'dogfriendly': dogfriendly, 'petfriendly': petfriendly, 'childfriendly': childfriendly,'gardenreq': gardenreq,
             'hug': hug,'training': training,'BCD': BCD,'score': score,'description': Description, 'image': image, 'summary': summary,'link': links})
        self.__recommender_data = pd.concat([self.__recommender_data, new_row.to_frame().T], ignore_index=True)

    def add_points(self, id, score):

        for row, dog in self.__recommender_data.iterrows():
            if dog['dog_id'] == id:
                self.__recommender_data.loc[row, 'score']+= score

    def make_recommendation(self, inputdata):

        for vraagnum, data in inputdata.iterrows():

            input = data['antwoorden']
            gewicht = data['gewichten']

            for row, dog in self.give_data().iterrows():

                # VRAAG 1: Welk geslacht prefereert u?  (eigenschap: geslacht --> reu, teef)
                if vraagnum == 0:
                    if dog['sex'] == input:
                        self.add_points(dog['dog_id'], 4.5 + ((gewicht - 50) / 50))

                # VRAAG 2: Welk grootte prefereert u?  (eigenschap: geslacht --> reu, teef)
                if vraagnum == 1:
                    if dog['size'] == input:
                        self.add_points(dog['dog_id'], 1 + ((gewicht - 50) / 50))

                # VRAAG 3: Leeftijd?  (eigenschap: geslacht --> reu, teef)
                if vraagnum == 2:
                    if dog['age'] == input:
                        self.add_points(dog['dog_id'], 1 + ((gewicht - 50) / 50))

                #VRAAG 4: Heeft u reeds (een) andere hond(en)? (eigenschap: ja, reu, teef, nee)

                if vraagnum == 3:
                    if dog['dogfriendly'] == input and input == 'Ja, een reu':
                        self.add_points(dog['dog_id'], 2 + ((gewicht - 50) / 50))

                    if dog['dogfriendly'] == input and input == 'Ja, een teef':
                        self.add_points(dog['dog_id'], 2 + ((gewicht - 50) / 50))

                    if dog['dogfriendly'] == input and input == 'Ja, meerdere honden':
                        self.add_points(dog['dog_id'], 2 + ((gewicht - 50) / 50))

                    if dog['dogfriendly'] == input and input == 'Nee':
                        self.add_points(dog['dog_id'], 2 + ((gewicht - 50) / 50))

                # VRAAG 5 want deze vragen hebben dezelfde puntenlogica en 1/1 matching met de dataset
                if vraagnum == 4:
                    if dog['petfriendly'] == input:
                        self.add_points(dog['dog_id'], 2 + ((gewicht - 50) / 50))

#-------- Vanaf hier speciale logica vragen:

                # VRAAG 6: Wat voor gezinssituatie heeft u? (eigenschap: ja, grote, nee)
                if vraagnum == 5:
                    if dog['childfriendly'] == input and input == "Enkel kinderen boven 12 jaar":
                        self.add_points(dog['dog_id'], 2 + ((gewicht - 50) / 50))

                    else:
                        self.give_all()

                # VRAAG 7: Heeft u een omheinde tuin? (eigenschap: default, moet tuin: ja)
                if vraagnum == 6:
                    if dog['gardenreq'] == input and input == "Nee":
                        self.add_points(dog['dog_id'], 2 + ((gewicht - 50) / 50))

                    if dog['gardenreq'] == input and input == "Ja":
                        self.give_all()

                # VRAAG 8: Zoekt u een echte knuffelbeer? (eigenschap: default, ja)
                if vraagnum == 7:
                    if dog['hug'] == input and input == "Ja, knuffelkontjes!":  # V8
                        self.add_points(dog['dog_id'], 1 + ((gewicht - 50) / 50))
                    else:
                        self.give_all()

                # VRAAG 9: Zou u graag training en/of hondensport doen met de hond? (eigenschap: default, training/sport)
                if vraagnum == 8:
                    if dog['training'] == input and input == "Ja, graag!":  # V9
                        self.add_points(dog['dog_id'], 1 + ((gewicht - 50) / 50))
                    else:
                         self.give_all()

                #VRAAG 10: Prefereert u een hond die getraind wordt via het Belgian Cell Dogs project?* (eigenschap: default, BCDp)
                if vraagnum == 9:
                    if dog['BCD'] == input and input == "Ja":  # V10
                        self.add_points(dog['dog_id'], 1 + ((gewicht - 50) / 50))
                    else:
                         self.give_all()

        return self.give_top4()

    def give_all(self):
        for row, dog in self.__recommender_data.iterrows():
            self.__recommender_data.loc[row, 'score']+= 1/len(self.__recommender_data)

    def give_data(self):
        return self.__recommender_data

    def search_dupl(self):
        for row1, dog1 in self.__recommender_data.iterrows():
            count =1
            for row2, dog2 in self.__recommender_data.iterrows():
                if row1==row2:
                    break
                else:
                    if dog1['name']== dog2['name']:
                        self.__recommender_data.loc[row1, 'score'] = self.__recommender_data.loc[row2, 'score']

    def give_top4(self):

        self.search_dupl()
        for row, data in self.__recommender_data.iterrows():
            x = self.__recommender_data.loc[row, 'score']
            self.__recommender_data.loc[row, 'score'] = round(x, 3)

        lijst = []
        for row, data in self.__recommender_data.sort_values(by=['score', 'name'], ascending=False).head(4).iterrows():
            dicti = {}
            dicti["name"] = data['name']
            dicti["imgURL"] = data['image']
            dicti["description"] = data['description']
            dicti["summary"] = data['summary']
            dicti["link"] = data['link']
            lijst.append(dicti)
        return lijst

    def export_excel(self, lijst):
        workbook = load_workbook('/app/records.xlsx')
        ws = workbook.active
        identificatie= id(lijst)
        inzet = [identificatie]
        for dicts in lijst:
            for key, value in dicts.items():
                inzet.append(value)

        ws.append(inzet)
        workbook.template = False
        workbook.save('/app/records.xlsx')
        return identificatie

    def read_excel(self, id):
        workbook = load_workbook('/app/records.xlsx')
        ws = workbook.active
        lijst=[]
        for row in ws.iter_rows(min_row=2, max_col= 21):
            if row[0].value == id:
                for x in range(0,16):
                    if x==0 or x==5 or x==10 or x==15:

                        dicti={}
                        dicti["name"] = row[1+x].value
                        dicti["imgURL"] = row[2+x].value
                        dicti["description"] = row[3+x].value
                        dicti["summary"] = row[4+x].value
                        dicti["link"] = row[5+x].value
                        lijst.append(dicti)
                        lijst
        return lijst

    def find_and_export(self, id, lijst):
        workbook = load_workbook('/app/records.xlsx')
        ws = workbook.active
        for index in ws.iter_rows(min_row=2):
            if index[0].value == id:

                for column in ["Y", "Z", "AA", "AB"]:
                    if column == "Y":
                        cell_name = "{}{}".format(column, index[0].row)
                        ws[cell_name].value  = lijst[0]
                    if column == "Z":
                        cell_name = "{}{}".format(column, index[0].row)
                        ws[cell_name].value = lijst[1]
                    if column == "AA":
                        cell_name = "{}{}".format(column, index[0].row)
                        ws[cell_name].value = lijst[2]
                    if column == "AB":
                        cell_name = "{}{}".format(column, index[0].row)
                        ws[cell_name].value = lijst[3]

        workbook.save('/app/records.xlsx')

